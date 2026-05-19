"""
core/optimizer.py
=================
Módulo de parseo, validación y asignación optimizada de horarios docentes.
No depende de Streamlit ni de ningún framework web.
"""

from __future__ import annotations

import io
import unicodedata
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONSTANTES
# ---------------------------------------------------------------------------
DIAS_COLS: dict[str, tuple[int, int]] = {
    "LUNES":     (12, 24),
    "MARTES":    (25, 37),
    "MIERCOLES": (38, 50),
    "JUEVES":    (51, 63),
    "VIERNES":   (64, 76),
    "SABADO":    (77, 89),
    "DOMINGO":   (90, 102),
}
SLOT_HOURS = [7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]

# Deptos HCM → lengua que puede enseñar
DEPTO_LENGUA: dict[str, str | None] = {
    "INGLES":                      "INGLÉS",
    "EDUCONTINUA ELT":             "INGLÉS",
    "FRANCES":                     "FRANCÉS",
    "ESPANOL":                     "ESPAÑOL",
    "VIVE_ESPANOL":                "ESPAÑOL",
    "ADMON LENGUAS":               None,
    "GESTION":                     None,
    "PEDAGOGIA DE LOS SABERES":    None,
    "PROYECTOS INNOVACION":        None,
    "ASESORIAS FAC COMUNICACION":  None,
    "PLAN DE FORMACION EMPLEADOS": None,
    "OTRAS LENGUAS PORTUG MANDA":  None,
    "IDIOMAS SABANA":              None,
}

DIA_ORDEN = {
    "LUNES": 1, "MARTES": 2, "MIERCOLES": 3, "JUEVES": 4,
    "VIERNES": 5, "SABADO": 6, "DOMINGO": 7,
}

# ---------------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------------

def normalize(text: str) -> str:
    """Minúsculas, sin tildes, sin espacios extra."""
    if not isinstance(text, str):
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFD", text)
    return "".join(c for c in text if unicodedata.category(c) != "Mn")


def normalize_day(day: str) -> str:
    """Devuelve el día normalizado en MAYÚSCULAS sin tilde."""
    return normalize(day).upper()


def parse_time(value: Any) -> int | None:
    """Convierte hora a entero. Acepta '08:00', '8:00', datetime.time, etc."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip()
    if ":" in s:
        return int(s.split(":")[0])
    if "." in s:
        return int(s.split(".")[0])
    try:
        return int(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# CARGA Y VALIDACIÓN
# ---------------------------------------------------------------------------

def load_excel(file) -> tuple[pd.DataFrame | None, pd.DataFrame | None, str | None]:
    """
    Carga el Excel y devuelve (df_disp_raw, df_hor, error).
    error es None si todo está correcto.
    """
    try:
        xl = pd.ExcelFile(file)
    except Exception as e:
        return None, None, f"No se pudo leer el archivo Excel: {e}"

    sheets_lower = {s.lower(): s for s in xl.sheet_names}

    disp_key = next((k for k in sheets_lower if "disponib" in k), None)
    hor_key  = next((k for k in sheets_lower if "horario" in k or "fijo" in k), None)

    if not disp_key:
        return None, None, "No se encontró la hoja de Disponibilidad en el archivo."
    if not hor_key:
        return None, None, "No se encontró la hoja de Horarios Fijos en el archivo."

    df_disp_raw = pd.read_excel(file, sheet_name=sheets_lower[disp_key], header=None)

    df_hor = pd.read_excel(file, sheet_name=sheets_lower[hor_key])
    # Normalizar nombres de columna
    df_hor.columns = [str(c).lower().strip() for c in df_hor.columns]
    # Eliminar filas sin id_horario
    id_col = "id_horario" if "id_horario" in df_hor.columns else df_hor.columns[0]
    df_hor = df_hor.dropna(subset=[id_col]).reset_index(drop=True)

    err = validate_horarios(df_hor)
    if err:
        return None, None, err

    return df_disp_raw, df_hor, None


def validate_horarios(df_hor: pd.DataFrame) -> str | None:
    required = {"lengua", "curso", "grupo", "dia", "hora_inicio", "hora_fin"}
    missing = required - set(df_hor.columns)
    if missing:
        return f"Faltan columnas en Horarios Fijos: {', '.join(sorted(missing))}"
    return None


# ---------------------------------------------------------------------------
# PARSEO DE DISPONIBILIDAD
# ---------------------------------------------------------------------------

def parse_disponibilidad(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Convierte la hoja cruda en registros: (profesor, depto, dia, hora, disponible).
    """
    data = df_raw.iloc[2:].reset_index(drop=True)
    records: list[dict] = []

    for _, row in data.iterrows():
        nombre = str(row[0]).strip() if pd.notna(row[0]) else ""
        if not nombre or nombre.lower() in ("nan", ""):
            continue
        depto = str(row[11]).strip() if pd.notna(row[11]) else ""

        for dia_norm, (col_start, col_end) in DIAS_COLS.items():
            for offset, hora in enumerate(SLOT_HOURS):
                col_idx = col_start + offset
                if col_idx > col_end:
                    break
                val = row[col_idx] if col_idx < len(row) else 0
                try:
                    disponible = int(float(val)) if pd.notna(val) else 0
                except (ValueError, TypeError):
                    disponible = 0
                records.append({
                    "profesor": nombre,
                    "depto": depto,
                    "dia": dia_norm,
                    "hora": hora,
                    "disponible": disponible,
                })

    return pd.DataFrame(records)


def get_professors_list(df_raw: pd.DataFrame) -> list[str]:
    """Retorna lista de nombres de profesores desde la hoja de disponibilidad."""
    data = df_raw.iloc[2:].reset_index(drop=True)
    return [
        str(row[0]).strip()
        for _, row in data.iterrows()
        if pd.notna(row[0]) and str(row[0]).strip().lower() not in ("nan", "")
    ]


def get_horarios_labels(df_hor: pd.DataFrame) -> list[dict]:
    """
    Retorna lista de dicts con {id, label} para el desplegable de clases.
    Formato: 'ID - Lengua - Curso - Grupo - Día - Hora inicio - Hora fin'
    """
    labels = []
    for _, row in df_hor.iterrows():
        id_h    = str(row.get("id_horario", "")).strip()
        lengua  = str(row.get("lengua", "")).strip()
        curso   = str(row.get("curso", "")).strip()
        grupo   = str(row.get("grupo", "")).strip()
        dia     = str(row.get("dia", "")).strip()
        h_ini   = str(row.get("hora_inicio", "")).strip()
        h_fin   = str(row.get("hora_fin", "")).strip()
        label = f"{id_h} · {lengua} · {curso} · Grupo {grupo} · {dia} {h_ini}-{h_fin}"
        labels.append({
            "id": id_h,
            "label": label,
            "curso": curso,
            "lengua": lengua,
        })
    return labels


# ---------------------------------------------------------------------------
# COMPATIBILIDAD LENGUA
# ---------------------------------------------------------------------------

def lengua_de_depto(depto: str) -> str | None:
    key = normalize(depto).upper()
    for k, v in DEPTO_LENGUA.items():
        if normalize(k).upper() == key:
            return v
    return None


def es_compatible(depto_prof: str, lengua_curso: str) -> bool:
    """True si el profesor puede enseñar esa lengua."""
    if not lengua_curso or lengua_curso.lower() in ("nan", "none", ""):
        return True  # sin lengua definida → cualquier prof sirve
    l = lengua_de_depto(depto_prof)
    if l is None:
        return True  # depto sin restricción
    return normalize(l) == normalize(lengua_curso)


# ---------------------------------------------------------------------------
# ASIGNACIÓN OPTIMIZADA (PuLP con fallback greedy)
# ---------------------------------------------------------------------------

def _greedy_assign(
    profesores: list[str],
    depto_map: dict[str, str],
    avail_idx: dict[tuple, int],
    horarios_sorted: list[dict],
    restricciones: list[dict],
) -> pd.DataFrame:
    """
    Asignación greedy mejorada con soporte de restricciones.
    Usada como fallback si PuLP no está disponible.
    """
    # Índices rápidos de restricciones
    rest_obligatoria: dict[str, str] = {}   # id_horario → profesor
    rest_exclusivo: dict[str, str]   = {}   # id_horario → único profesor permitido
    rest_solo_una: dict[str, str]    = {}   # profesor → único id_horario que puede tomar
    rest_permitida: dict[str, set]   = {}   # id_horario → set de profesores preferidos
    rest_prof_solo_curso: dict[str, str] = {}   # profesor → curso/nivel único
    rest_curso_solo_prof: dict[str, str] = {}   # curso/nivel → profesor único

    for r in restricciones:
        tipo = r.get("tipo", "")
        prof = r.get("profesor", "")
        id_h = r.get("id_horario", "")
        if tipo in ("obligatoria", "debe_dictar"):
            rest_obligatoria[id_h] = prof
        elif tipo == "exclusivo":
            rest_exclusivo[id_h] = prof
        elif tipo == "solo_una":
            rest_solo_una[prof] = id_h
        elif tipo == "permitida":
            rest_permitida.setdefault(id_h, set()).add(prof)
        elif tipo == "profesor_solo_nivel":
            curso = r.get("curso", "")
            if prof and curso:
                rest_prof_solo_curso[prof] = curso
        elif tipo == "nivel_solo_profesor":
            curso = r.get("curso", "")
            if prof and curso:
                rest_curso_solo_prof[normalize(curso)] = prof

    asignaciones_dia: dict[tuple, set] = {}
    ocupado: set[tuple] = set()
    prof_ya_asignado: dict[str, str] = {}  # para rest_solo_una: prof→id_h asignado

    resultados = []

    # Primero procesar obligatorias para bloquear slots
    obligatorias_procesadas = set()

    def _asignar(hor: dict, prof_elegido: str, obs: str) -> dict:
        dia   = normalize_day(str(hor.get("dia", "")))
        h_ini = parse_time(hor.get("hora_inicio"))
        h_fin = parse_time(hor.get("hora_fin"))
        horas = list(range(h_ini, h_fin)) if h_ini and h_fin else []
        k_dia = (prof_elegido, dia)
        if k_dia not in asignaciones_dia:
            asignaciones_dia[k_dia] = set()
        for h in horas:
            asignaciones_dia[k_dia].add(h)
            ocupado.add((prof_elegido, dia, h))
        return {
            "ID Horario":       str(hor.get("id_horario", "")),
            "Semestre":         str(hor.get("semestre", "")),
            "Lengua":           str(hor.get("lengua", "")),
            "Tipo curso":       str(hor.get("tipo_curso", "")),
            "Curso":            str(hor.get("curso", "")),
            "Grupo":            str(hor.get("grupo", "")),
            "Día":              str(hor.get("dia", "")),
            "Hora inicio":      str(hor.get("hora_inicio", "")),
            "Hora fin":         str(hor.get("hora_fin", "")),
            "Modalidad":        str(hor.get("modalidad", "")),
            "Profesor asignado": prof_elegido,
            "Estado":           "Asignado",
            "Observaciones":    obs,
        }

    def _no_asignar(hor: dict, obs: str) -> dict:
        return {
            "ID Horario":        str(hor.get("id_horario", "")),
            "Semestre":          str(hor.get("semestre", "")),
            "Lengua":            str(hor.get("lengua", "")),
            "Tipo curso":        str(hor.get("tipo_curso", "")),
            "Curso":             str(hor.get("curso", "")),
            "Grupo":             str(hor.get("grupo", "")),
            "Día":               str(hor.get("dia", "")),
            "Hora inicio":       str(hor.get("hora_inicio", "")),
            "Hora fin":          str(hor.get("hora_fin", "")),
            "Modalidad":         str(hor.get("modalidad", "")),
            "Profesor asignado": "SIN ASIGNAR",
            "Estado":            "No asignado",
            "Observaciones":     obs,
        }

    for hor in horarios_sorted:
        id_h   = str(hor.get("id_horario", "")).strip()
        dia    = normalize_day(str(hor.get("dia", "")))
        h_ini  = parse_time(hor.get("hora_inicio"))
        h_fin  = parse_time(hor.get("hora_fin"))
        lengua = str(hor.get("lengua", "")).strip()
        curso_actual = str(hor.get("curso", "")).strip()

        if not dia or h_ini is None or h_fin is None:
            resultados.append(_no_asignar(hor, "Datos de horario incompletos o inválidos."))
            continue

        horas_bloque = list(range(h_ini, h_fin))

        # Verificar restricciones de horario específico (tipo D)
        hora_override = None
        for r in restricciones:
            if r.get("tipo") == "horario_especifico" and r.get("id_horario") == id_h:
                dia_r  = normalize_day(r.get("dia_especifico", ""))
                ini_r  = parse_time(r.get("hora_inicio_especifica"))
                fin_r  = parse_time(r.get("hora_fin_especifica"))
                if dia_r and ini_r and fin_r:
                    hora_override = (dia_r, ini_r, fin_r)
                break

        if hora_override:
            dia, h_ini, h_fin = hora_override
            horas_bloque = list(range(h_ini, h_fin))

        # ── Restricción obligatoria / debe_dictar ──
        if id_h in rest_obligatoria:
            prof_req = rest_obligatoria[id_h]
            depto_req = depto_map.get(prof_req, "")
            # Verificar disponibilidad
            ok = all(
                avail_idx.get((prof_req, dia, h), 0) == 1 and (prof_req, dia, h) not in ocupado
                for h in horas_bloque
            )
            if not ok:
                obs = (f"Restricción obligatoria: {prof_req} no está disponible o tiene "
                       "conflicto en este horario.")
                resultados.append(_no_asignar(hor, obs))
            else:
                obs = f"Asignado por restricción obligatoria."
                resultados.append(_asignar(hor, prof_req, obs))
                if prof_req in rest_solo_una:
                    prof_ya_asignado[prof_req] = id_h
            continue

        # ── Restricción exclusivo ──
        if id_h in rest_exclusivo:
            prof_excl = rest_exclusivo[id_h]
            ok = all(
                avail_idx.get((prof_excl, dia, h), 0) == 1 and (prof_excl, dia, h) not in ocupado
                for h in horas_bloque
            )
            if not ok:
                obs = (f"Restricción exclusivo: solo {prof_excl} puede dictar esta clase "
                       "pero no está disponible o tiene conflicto.")
                resultados.append(_no_asignar(hor, obs))
            else:
                obs = f"Asignado por restricción de exclusividad."
                resultados.append(_asignar(hor, prof_excl, obs))
            continue

        # ── Candidatos normales ──
        candidatos: list[tuple[int, str]] = []
        permitidos = rest_permitida.get(id_h)  # None si no hay restricción de este tipo

        for prof in profesores:
            # Restricción: este profesor solo puede dictar un curso/nivel específico
            if prof in rest_prof_solo_curso:
                if normalize(curso_actual) != normalize(rest_prof_solo_curso[prof]):
                    continue

            # Restricción: este curso/nivel solo puede ser dictado por un profesor específico
            prof_unico_nivel = rest_curso_solo_prof.get(normalize(curso_actual))
            if prof_unico_nivel and prof != prof_unico_nivel:
                continue
            # Restricción solo_una: si el prof ya fue asignado a otra clase, saltar
            if prof in rest_solo_una and rest_solo_una[prof] != id_h:
                if prof in prof_ya_asignado:
                    continue

            depto = depto_map.get(prof, "")
            if not es_compatible(depto, lengua):
                continue

            # Si hay lista permitida, solo esos profesores pueden tomar esta clase
            if permitidos and prof not in permitidos:
                continue

            ok = all(
                avail_idx.get((prof, dia, h), 0) == 1 and (prof, dia, h) not in ocupado
                for h in horas_bloque
            )
            if ok:
                carga = len(asignaciones_dia.get((prof, dia), set()))
                candidatos.append((carga, prof))

        if not candidatos:
            # Determinar motivo
            compatibles = [
                p for p in profesores
                if es_compatible(depto_map.get(p, ""), lengua)
            ]
            if not compatibles:
                obs = f"No hay profesores registrados para la lengua '{lengua}'."
            else:
                todos_ocupados = all(
                    any((p, dia, h) in ocupado for h in horas_bloque)
                    for p in compatibles
                )
                if todos_ocupados:
                    obs = "Todos los profesores compatibles tienen conflicto de horario en este bloque."
                else:
                    obs = "No hay profesores disponibles para este día y horario."
            resultados.append(_no_asignar(hor, obs))
            continue

        # Elegir el de menor carga (balanceo)
        candidatos.sort()
        _, prof_elegido = candidatos[0]

        obs = ""
        if permitidos and prof_elegido in permitidos:
            obs = "Asignado según restricción de profesores permitidos."

        result_row = _asignar(hor, prof_elegido, obs)
        resultados.append(result_row)

        if prof_elegido in rest_solo_una:
            prof_ya_asignado[prof_elegido] = id_h

    return pd.DataFrame(resultados)


def _pulp_assign(
    profesores: list[str],
    depto_map: dict[str, str],
    avail_idx: dict[tuple, int],
    horarios_sorted: list[dict],
    restricciones: list[dict],
) -> pd.DataFrame | None:
    """
    Asignación mediante programación lineal entera con PuLP.
    Maximiza clases asignadas y balancea carga.
    Retorna None si PuLP no está disponible.
    """
    try:
        import pulp  # type: ignore
    except ImportError:
        return None

    # Índices de restricciones
    rest_obligatoria: dict[str, str] = {}
    rest_exclusivo: dict[str, str]   = {}
    rest_solo_una: dict[str, str]    = {}
    rest_no_permitida: dict[str, set] = {}  # id_horario → profs excluidos
    rest_permitida: dict[str, set]   = {}
    rest_prof_solo_curso: dict[str, str] = {}
    rest_curso_solo_prof: dict[str, str] = {}

    for r in restricciones:
        tipo = r.get("tipo", "")
        prof = r.get("profesor", "")
        id_h = r.get("id_horario", "")
        if tipo in ("obligatoria", "debe_dictar"):
            rest_obligatoria[id_h] = prof
        elif tipo == "exclusivo":
            rest_exclusivo[id_h] = prof
        elif tipo == "solo_una":
            rest_solo_una[prof] = id_h
        elif tipo == "permitida":
            rest_permitida.setdefault(id_h, set()).add(prof)
        elif tipo == "profesor_solo_nivel":
            curso = r.get("curso", "")
            if prof and curso:
                rest_prof_solo_curso[prof] = curso
        elif tipo == "nivel_solo_profesor":
            curso = r.get("curso", "")
            if prof and curso:
                rest_curso_solo_prof[normalize(curso)] = prof

    # Para restricciones exclusivas → todos los demás están excluidos
    for id_h, prof_unico in rest_exclusivo.items():
        excluidos = set(profesores) - {prof_unico}
        rest_no_permitida[id_h] = excluidos

    prob = pulp.LpProblem("HorariosDocentes", pulp.LpMaximize)

    # Variables: x[p][h] = 1 si el profesor p dicta el horario h
    horario_ids = [str(h.get("id_horario", "")) for h in horarios_sorted]
    valid_pairs: dict[tuple[str, str], bool] = {}

    for hor in horarios_sorted:
        id_h   = str(hor.get("id_horario", ""))
        dia    = normalize_day(str(hor.get("dia", "")))
        h_ini  = parse_time(hor.get("hora_inicio"))
        h_fin  = parse_time(hor.get("hora_fin"))
        lengua = str(hor.get("lengua", "")).strip()
        curso_actual = str(hor.get("curso", "")).strip()

        if not dia or h_ini is None or h_fin is None:
            continue
        horas_bloque = list(range(h_ini, h_fin))

        for prof in profesores:
            depto = depto_map.get(prof, "")
            if prof in rest_prof_solo_curso:
                if normalize(curso_actual) != normalize(rest_prof_solo_curso[prof]):
                    continue

            prof_unico_nivel = rest_curso_solo_prof.get(normalize(curso_actual))
            if prof_unico_nivel and prof != prof_unico_nivel:
                continue
            if not es_compatible(depto, lengua):
                continue
            # Restricción exclusivo / permitida
            if id_h in rest_no_permitida and prof in rest_no_permitida[id_h]:
                continue
            if id_h in rest_permitida and prof not in rest_permitida[id_h]:
                # Solo bloquear si hay lista de permitidos
                pass
            ok = all(avail_idx.get((prof, dia, h), 0) == 1 for h in horas_bloque)
            if ok:
                valid_pairs[(prof, id_h)] = True

    x = {
        (p, h): pulp.LpVariable(f"x_{p.replace(' ', '_')}_{h.replace('-', '_')}", cat="Binary")
        for (p, h) in valid_pairs
    }
        # ------------------------------------------------------------------
    # Objetivo:
    # 1. Maximizar cantidad de clases asignadas.
    # 2. Penalizar que un profesor tenga demasiados niveles distintos.
    # 3. Penalizar suavemente cargas demasiado altas.
    # ------------------------------------------------------------------

    asignados_total = pulp.lpSum(x.values())

    # Cursos/niveles existentes
    cursos = sorted({
        str(hor.get("curso", "")).strip()
        for hor in horarios_sorted
        if str(hor.get("curso", "")).strip()
    })

    # y[p, curso] = 1 si el profesor p dicta al menos una clase de ese curso/nivel
    y = {}

    for prof in profesores:
        for curso in cursos:
            y[(prof, curso)] = pulp.LpVariable(
                f"y_{prof.replace(' ', '_')}_{curso.replace(' ', '_')}",
                cat="Binary"
            )

    # Relacionar x con y:
    # si x[p, horario] = 1, entonces y[p, curso_del_horario] debe ser 1
    for hor in horarios_sorted:
        id_h = str(hor.get("id_horario", ""))
        curso = str(hor.get("curso", "")).strip()

        if not curso:
            continue

        for prof in profesores:
            if (prof, id_h) in x:
                prob += x[(prof, id_h)] <= y[(prof, curso)]

    # Número total de combinaciones profesor-nivel usadas
    # Mientras menor sea, más concentrados quedan los niveles por profesor
    niveles_distintos_total = pulp.lpSum(y.values())

    # Horas por profesor
    total_horas_por_prof = {
        prof: pulp.lpSum(
            x[(prof, id_h)] * (
                (parse_time(hor.get("hora_fin")) or 0)
                - (parse_time(hor.get("hora_inicio")) or 0)
            )
            for hor in horarios_sorted
            for id_h in [str(hor.get("id_horario", ""))]
            if (prof, id_h) in x
        )
        for prof in profesores
    }

    horas_totales = pulp.lpSum(total_horas_por_prof.values())

    # Función objetivo:
    # - 1000: asignar clases es lo más importante.
    # - 15: penaliza abrir demasiados niveles distintos por profesor.
    # - 0.01: penalización suave por horas, sirve como desempate.
    prob += (
        asignados_total * 1000
        - niveles_distintos_total * 15
        - horas_totales * 0.01
    )
    # Restricción: cada horario asignado a máximo 1 profesor
    for id_h in horario_ids:
        vars_h = [x[(p, id_h)] for p in profesores if (p, id_h) in x]
        if vars_h:
            prob += pulp.lpSum(vars_h) <= 1
        # ------------------------------------------------------------------
    # Restricción: mantener el mismo profesor para todas las sesiones
    # del mismo grupo académico.
    #
    # Ejemplo:
    # INGLÉS 1 - Grupo 2 - Miércoles
    # INGLÉS 1 - Grupo 2 - Jueves
    # deben quedar con el mismo profesor.
    # ------------------------------------------------------------------

    grupos_academicos = {}

    for hor in horarios_sorted:
        id_h = str(hor.get("id_horario", "")).strip()

        clave_grupo = (
            str(hor.get("lengua", "")).strip(),
            str(hor.get("curso", "")).strip(),
            str(hor.get("nivel", "")).strip(),
            str(hor.get("grupo", "")).strip(),
        )

        grupos_academicos.setdefault(clave_grupo, []).append(id_h)

    for clave_grupo, ids_grupo in grupos_academicos.items():
        if len(ids_grupo) <= 1:
            continue

        for prof in profesores:
            ids_validos_prof = [
                id_h for id_h in ids_grupo
                if (prof, id_h) in x
            ]

            if len(ids_validos_prof) <= 1:
                continue

            primer_id = ids_validos_prof[0]

            for otro_id in ids_validos_prof[1:]:
                prob += x[(prof, primer_id)] == x[(prof, otro_id)]
    # Restricción: sin solapamiento de horarios por profesor
    for prof in profesores:
        dia_horas: dict[tuple, list] = {}
        for hor in horarios_sorted:
            id_h   = str(hor.get("id_horario", ""))
            dia    = normalize_day(str(hor.get("dia", "")))
            h_ini  = parse_time(hor.get("hora_inicio"))
            h_fin  = parse_time(hor.get("hora_fin"))
            if h_ini is None or h_fin is None:
                continue
            if (prof, id_h) not in x:
                continue
            for h in range(h_ini, h_fin):
                key = (dia, h)
                dia_horas.setdefault(key, []).append(x[(prof, id_h)])
        for key, vars_slot in dia_horas.items():
            if len(vars_slot) > 1:
                prob += pulp.lpSum(vars_slot) <= 1

    # Restricciones obligatorias
    for id_h, prof_req in rest_obligatoria.items():
        if (prof_req, id_h) in x:
            prob += x[(prof_req, id_h)] == 1
        # Si no está en x → infactible, se maneja en post-proceso

    # Restricción solo_una
    for prof, id_h_unico in rest_solo_una.items():
        vars_otros = [x[(prof, id_h)] for id_h in horario_ids if (prof, id_h) in x and id_h != id_h_unico]
        if vars_otros:
            prob += pulp.lpSum(vars_otros) == 0

    # Resolver
    solver = pulp.PULP_CBC_CMD(msg=0, timeLimit=30)
    prob.solve(solver)

    if pulp.LpStatus[prob.status] not in ("Optimal", "Feasible"):
        return None  # Fallback a greedy

    # Construir resultado
    asignacion: dict[str, str] = {}
    for (prof, id_h), var in x.items():
        if pulp.value(var) and pulp.value(var) > 0.5:
            asignacion[id_h] = prof

    resultados = []
    for hor in horarios_sorted:
        id_h   = str(hor.get("id_horario", ""))
        dia    = str(hor.get("dia", ""))
        h_ini  = parse_time(hor.get("hora_inicio"))
        h_fin  = parse_time(hor.get("hora_fin"))

        base = {
            "ID Horario":  id_h,
            "Semestre":    str(hor.get("semestre", "")),
            "Lengua":      str(hor.get("lengua", "")),
            "Tipo curso":  str(hor.get("tipo_curso", "")),
            "Curso":       str(hor.get("curso", "")),
            "Grupo":       str(hor.get("grupo", "")),
            "Día":         dia,
            "Hora inicio": str(hor.get("hora_inicio", "")),
            "Hora fin":    str(hor.get("hora_fin", "")),
            "Modalidad":   str(hor.get("modalidad", "")),
        }

        if not dia or h_ini is None or h_fin is None:
            resultados.append({**base, "Profesor asignado": "SIN ASIGNAR", "Estado": "No asignado",
                                "Observaciones": "Datos de horario incompletos."})
            continue

        if id_h in asignacion:
            obs = ""
            if id_h in rest_obligatoria:
                obs = "Asignado por restricción obligatoria."
            elif id_h in rest_exclusivo:
                obs = "Asignado por restricción de exclusividad."
            elif id_h in rest_permitida:
                obs = "Asignado según restricción de profesores permitidos."
            resultados.append({**base, "Profesor asignado": asignacion[id_h],
                                "Estado": "Asignado", "Observaciones": obs})
        else:
            # Determinar motivo
            dia_n  = normalize_day(dia)
            horas  = list(range(h_ini, h_fin))
            lengua = str(hor.get("lengua", ""))
            compatibles = [p for p in profesores if es_compatible(depto_map.get(p, ""), lengua)]
            if not compatibles:
                obs = f"No hay profesores registrados para la lengua '{lengua}'."
            elif id_h in rest_obligatoria:
                obs = (f"Restricción obligatoria: {rest_obligatoria[id_h]} no disponible "
                       "o tiene conflicto.")
            elif id_h in rest_exclusivo:
                obs = (f"Restricción exclusivo: solo {rest_exclusivo[id_h]} puede dictar "
                       "esta clase pero no está disponible o tiene conflicto.")
            else:
                any_avail = any(
                    all(avail_idx.get((p, dia_n, h), 0) == 1 for h in horas)
                    for p in compatibles
                )
                if any_avail:
                    obs = "Todos los profesores disponibles tienen conflicto de horario."
                else:
                    obs = "No hay profesores disponibles para este día y horario."
            resultados.append({**base, "Profesor asignado": "SIN ASIGNAR",
                                "Estado": "No asignado", "Observaciones": obs})

    return pd.DataFrame(resultados)


def optimize_assignments(
    df_disp: pd.DataFrame,
    df_hor: pd.DataFrame,
    restricciones: list[dict] | None = None,
) -> tuple[pd.DataFrame, str]:
    """
    Punto de entrada principal del optimizador.
    Intenta PuLP primero; si no está disponible usa greedy.
    Retorna (df_resultado, metodo_usado).
    """
    if restricciones is None:
        restricciones = []

    avail_idx = {
        (row["profesor"], row["dia"], row["hora"]): row["disponible"]
        for _, row in df_disp.iterrows()
    }

    profesores = list(df_disp["profesor"].unique())
    depto_map  = {
        row["profesor"]: row["depto"]
        for _, row in df_disp.drop_duplicates("profesor").iterrows()
    }

    # Ordenar horarios
    horarios_sorted = sorted(
        df_hor.to_dict("records"),
        key=lambda r: (
            DIA_ORDEN.get(normalize_day(str(r.get("dia", ""))), 8),
            parse_time(r.get("hora_inicio")) or 0,
        ),
    )

    # Intentar PuLP
    df_result = _pulp_assign(profesores, depto_map, avail_idx, horarios_sorted, restricciones)
    if df_result is not None:
        return df_result, "PuLP (optimización lineal entera)"

    # Fallback greedy
    df_result = _greedy_assign(profesores, depto_map, avail_idx, horarios_sorted, restricciones)
    return df_result, "Greedy mejorado (PuLP no disponible)"


# ---------------------------------------------------------------------------
# GENERACIÓN DEL EXCEL DE SALIDA
# ---------------------------------------------------------------------------

def build_carga_por_nivel(df_result: pd.DataFrame) -> pd.DataFrame:
    """
    Construye una tabla tipo matriz para visualizar la carga por profesor.

    Cuenta grupos únicos por profesor y curso/nivel, no bloques horarios.
    Por ejemplo, si INGLÉS 1 Grupo 2 tiene clase miércoles y jueves,
    se cuenta como 1 grupo, no como 2.
    """
    df_asig = df_result[df_result["Estado"] == "Asignado"].copy()

    if df_asig.empty:
        return pd.DataFrame(columns=[
            "NOMBRE PROFESOR",
            "NIVEL 1",
            "# grupos 1",
            "TOTAL GRUPOS",
        ])

    df_asig["grupo_unico"] = (
        df_asig["Lengua"].astype(str).str.strip()
        + " | "
        + df_asig["Curso"].astype(str).str.strip()
        + " | "
        + df_asig["Grupo"].astype(str).str.strip()
    )

    carga = (
        df_asig
        .drop_duplicates(subset=["Profesor asignado", "Curso", "grupo_unico"])
        .groupby(["Profesor asignado", "Curso"])
        .size()
        .reset_index(name="# grupos")
        .sort_values(
            ["Profesor asignado", "# grupos"],
            ascending=[True, False]
        )
    )

    rows = []

    for profesor, grupo_prof in carga.groupby("Profesor asignado"):
        niveles = grupo_prof[["Curso", "# grupos"]].values.tolist()

        row = {
            "NOMBRE PROFESOR": profesor,
            "TOTAL GRUPOS": int(grupo_prof["# grupos"].sum()),
        }

        for i, item in enumerate(niveles, start=1):
            curso = item[0]
            num_grupos = int(item[1])

            row[f"NIVEL {i}"] = curso
            row[f"# grupos {i}"] = num_grupos

        rows.append(row)

    df_carga = pd.DataFrame(rows)

    max_niveles = 0

    for col in df_carga.columns:
        if col.startswith("NIVEL "):
            try:
                numero = int(col.replace("NIVEL ", ""))
                max_niveles = max(max_niveles, numero)
            except ValueError:
                pass

    ordered_cols = ["NOMBRE PROFESOR"]

    for i in range(1, max_niveles + 1):
        ordered_cols.append(f"NIVEL {i}")
        ordered_cols.append(f"# grupos {i}")

    ordered_cols.append("TOTAL GRUPOS")

    for col in ordered_cols:
        if col not in df_carga.columns:
            df_carga[col] = ""

    df_carga = df_carga[ordered_cols]
    df_carga = df_carga.sort_values("TOTAL GRUPOS", ascending=False)

    return df_carga


def _style_sheet(ws, header_color: str, alt_color: str) -> None:
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", start_color=header_color)
    a_fill = PatternFill("solid", start_color=alt_color)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = a_fill if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if fill.fill_type:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    ws.row_dimensions[1].height = 30


def generate_excel(df_result: pd.DataFrame, metodo: str = "") -> bytes:
    """Genera el Excel de salida y lo retorna como bytes."""
    wb = Workbook()
    ws_asig = wb.active
    ws_asig.title = "Asignación"

    # ── Hoja Asignación ──
    cols = df_result.columns.tolist()
    ws_asig.append(cols)

    red_fill = PatternFill("solid", start_color="FFCCCC")
    green_fill = PatternFill("solid", start_color="C6EFCE")
    estado_idx = cols.index("Estado")

    for _, row in df_result.iterrows():
        ws_asig.append(row.tolist())

    for row in ws_asig.iter_rows(min_row=2, max_row=ws_asig.max_row):
        estado_cell = row[estado_idx]
        estado_cell.fill = green_fill if estado_cell.value == "Asignado" else red_fill

    _style_sheet(ws_asig, "1F4E79", "D6E4F0")
    ws_asig.freeze_panes = "A2"

    # ── Hoja Resumen ──
    ws_res = wb.create_sheet("Resumen")

    total = len(df_result)
    asignados = int((df_result["Estado"] == "Asignado").sum())
    no_asig = total - asignados
    cobertura = f"{round(asignados / total * 100, 1)}%" if total else "0%"

    ws_res.append(["Métrica", "Valor"])
    for label, val in [
        ("Total horarios procesados", total),
        ("Total asignados", asignados),
        ("Total sin asignar", int(no_asig)),
        ("Cobertura", cobertura),
        ("Método de optimización", metodo),
    ]:
        ws_res.append([label, val])

    ws_res.append([])
    ws_res.append(["Carga por Profesor", "Bloques asignados", "Horas estimadas"])

    carga = (
        df_result[df_result["Estado"] == "Asignado"]
        .groupby("Profesor asignado")
        .apply(lambda g: pd.Series({
            "Bloques": len(g),
            "Horas": sum(
                (parse_time(str(f)) or 0) - (parse_time(str(i)) or 0)
                for i, f in zip(g["Hora inicio"], g["Hora fin"])
            ),
        }))
        .reset_index()
        .sort_values("Horas", ascending=False)
    )

    for _, r in carga.iterrows():
        ws_res.append([
            r["Profesor asignado"],
            int(r["Bloques"]),
            int(r["Horas"])
        ])

    _style_sheet(ws_res, "375623", "E2EFDA")

    # ── Hoja Carga por nivel ──
    ws_nivel = wb.create_sheet("Carga por nivel")

    df_carga_nivel = build_carga_por_nivel(df_result)

    ws_nivel.append(df_carga_nivel.columns.tolist())

    for _, row in df_carga_nivel.iterrows():
        ws_nivel.append(row.tolist())

    _style_sheet(ws_nivel, "1F4E79", "D6E4F0")
    ws_nivel.freeze_panes = "A2"

    # Resaltar TOTAL GRUPOS en rojo
    if "TOTAL GRUPOS" in df_carga_nivel.columns:
        total_col_idx = df_carga_nivel.columns.tolist().index("TOTAL GRUPOS") + 1

        for row in ws_nivel.iter_rows(min_row=2, max_row=ws_nivel.max_row):
            total_cell = row[total_col_idx - 1]
            total_cell.font = Font(
                bold=True,
                color="FF0000",
                name="Arial",
                size=10
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

